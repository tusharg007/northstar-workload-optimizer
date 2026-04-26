"""
North Star — ROI Excel Generator
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "roi_calculation.xlsx")

def create_roi_workbook():
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    gold_fill = PatternFill(start_color='FFD23F', end_color='FFD23F', fill_type='solid')
    green_fill = PatternFill(start_color='2ECC71', end_color='2ECC71', fill_type='solid')
    red_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
    bold = Font(name='Calibri', bold=True, size=11)
    curr = '#,##0'
    pct = '0.0%'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    # ── Sheet 1: Current State Costs ──
    ws1 = wb.active
    ws1.title = "Current State"
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 20

    ws1['A1'] = "CURRENT STATE — EXPENSE PROCESSING COSTS"
    ws1['A1'].font = Font(name='Calibri', bold=True, size=14, color='1B2A4A')
    
    headers = ['Cost Element', 'Volume/Month', 'Time (min)', 'Hourly Rate ($)', 'Monthly Cost ($)', 'Annual Cost ($)']
    for i, h in enumerate(headers, 1):
        ws1.cell(row=3, column=i, value=h)
    style_header(ws1, 3, 6)

    data = [
        ['Employee Data Entry', 420, 33, 45, '=B4*C4/60*D4', '=E4*12'],
        ['Manager Review & Approval', 420, 37, 75, '=B5*C5/60*D5', '=E5*12'],
        ['Finance Processing', 420, 40, 55, '=B6*C6/60*D6', '=E6*12'],
        ['Rework/Error Correction (18%)', '=B4*0.18', 30, 55, '=B7*C7/60*D7', '=E7*12'],
        ['Late Payment Penalties', '', '', '', '=1058543*0.05/12', '=E8*12'],
    ]
    for r, row_data in enumerate(data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 5:
                cell.number_format = curr

    ws1.cell(row=9, column=1, value="TOTAL ANNUAL COST").font = bold
    ws1.cell(row=9, column=6, value='=SUM(F4:F8)').font = bold
    ws1.cell(row=9, column=6).number_format = curr
    ws1.cell(row=9, column=6).fill = red_fill
    ws1.cell(row=9, column=6).font = Font(bold=True, color='FFFFFF', size=12)

    # ── Sheet 2: Future State ──
    ws2 = wb.create_sheet("Future State")
    ws2.column_dimensions['A'].width = 35
    for c in 'BCDEF':
        ws2.column_dimensions[c].width = 18

    ws2['A1'] = "FUTURE STATE — AUTOMATED PROCESSING COSTS"
    ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color='1B2A4A')

    for i, h in enumerate(headers, 1):
        ws2.cell(row=3, column=i, value=h)
    style_header(ws2, 3, 6)

    future_data = [
        ['OCR + Auto-Entry', 420, 5, 45, '=B4*C4/60*D4', '=E4*12'],
        ['Auto-Route + Mobile Approve', 420, 8, 75, '=B5*C5/60*D5', '=E5*12'],
        ['Auto-Validate + Process', 420, 5, 55, '=B6*C6/60*D6', '=E6*12'],
        ['Rework (reduced to 5%)', '=B4*0.05', 15, 55, '=B7*C7/60*D7', '=E7*12'],
        ['Software Licensing (annual)', '', '', '', '=2500', '=E8*12'],
    ]
    for r, row_data in enumerate(future_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 5:
                cell.number_format = curr

    ws2.cell(row=9, column=1, value="TOTAL FUTURE COST").font = bold
    ws2.cell(row=9, column=6, value='=SUM(F4:F8)').font = bold
    ws2.cell(row=9, column=6).number_format = curr
    ws2.cell(row=9, column=6).fill = green_fill

    # ── Sheet 3: ROI Calculation ──
    ws3 = wb.create_sheet("ROI Analysis")
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 20

    ws3['A1'] = "ROI ANALYSIS — NORTH STAR WORKLOAD OPTIMIZER"
    ws3['A1'].font = Font(name='Calibri', bold=True, size=14, color='1B2A4A')

    roi_data = [
        ('Current Annual Cost', "='Current State'!F9", '', ''),
        ('Future Annual Cost', "='Future State'!F9", '', ''),
        ('Annual Savings', '=B3-B4', '', ''),
        ('', '', '', ''),
        ('Implementation Cost (one-time)', 45000, '', ''),
        ('Annual Software Cost', '=B4', '', ''),
        ('Total Year-1 Investment', '=B7+B8', '', ''),
        ('', '', '', ''),
        ('ROI % (Year 1)', '=(B5-B9)/B9', '', ''),
        ('ROI % (Year 2+)', '=(B5-B8)/B8', '', ''),
        ('Payback Period (months)', '=B9/B5*12', '', ''),
        ('', '', '', ''),
        ('', 'Conservative', 'Baseline', 'Optimistic'),
        ('Efficiency Gain', 0.60, 0.78, 0.85),
        ('Annual Savings', '=B3*B16', '=B3*C16', '=B3*D16'),
        ('ROI %', '=(B17-B9)/B9', '=(C17-B9)/B9', '=(D17-B9)/B9'),
    ]

    for r, (label, *vals) in enumerate(roi_data, 3):
        ws3.cell(row=r, column=1, value=label).font = bold if label else Font()
        for c, v in enumerate(vals, 2):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = thin_border
            if 'ROI' in label:
                cell.number_format = pct
            elif 'Gain' in label:
                cell.number_format = pct
            elif any(x in label for x in ['Cost', 'Savings', 'Investment']):
                cell.number_format = curr

    ws3.cell(row=5, column=1).fill = green_fill
    style_header(ws3, 15, 4)

    wb.save(OUTPUT)
    print(f"ROI workbook saved: {OUTPUT}")

if __name__ == "__main__":
    create_roi_workbook()
