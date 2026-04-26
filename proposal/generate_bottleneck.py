"""
North Star — Bottleneck Table Excel Generator
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bottleneck_table.xlsx")

def create():
    wb = openpyxl.Workbook()
    hf = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    red = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
    green = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
    bold = Font(name='Calibri', bold=True, size=11)
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

    def header(ws, row, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, Alignment(horizontal='center', wrap_text=True), border

    # Sheet 1: Process Steps
    ws1 = wb.active
    ws1.title = "Process Steps"
    for col in 'ABCDEFG':
        ws1.column_dimensions[col].width = 20
    ws1['A1'] = "CURRENT EXPENSE REPORTING PROCESS — TIME MOTION STUDY"
    ws1['A1'].font = Font(bold=True, size=14, color='1B2A4A')

    header(ws1, 3, ['Step #', 'Lane', 'Activity', 'Active Time (min)', 'Wait Time (hrs)', 'Total Time', 'Bottleneck?'])
    steps = [
        [1, 'Employee', 'Incur expense & collect receipt', 5, 0, '5 min', 'No'],
        [2, 'Employee', 'Log into expense portal', 8, 0, '8 min', 'No'],
        [3, 'Employee', 'Manual data entry', 15, 0, '15 min', 'YES'],
        [4, 'Employee', 'Scan/photograph receipts', 10, 0, '10 min', 'YES'],
        [5, 'Employee', 'Submit report via email', 5, 0, '5 min', 'No'],
        [6, 'System', 'Route submission to manager', 2, 48, '48 hrs', 'YES'],
        [7, 'Manager', 'Open & review report', 12, 0, '12 min', 'No'],
        [8, 'Manager', 'Cross-check against policy', 20, 0, '20 min', 'YES'],
        [9, 'Manager', 'Approve/reject & forward', 5, 24, '24 hrs', 'YES'],
        [10, 'Finance', 'Validate receipts & compliance', 25, 0, '25 min', 'No'],
        [11, 'Finance', 'Process reimbursement in ERP', 15, 72, '72 hrs', 'YES'],
        [12, 'System', 'Execute payment & notify', 3, 24, '24 hrs', 'No'],
    ]
    for r, row_data in enumerate(steps, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = border
            if row_data[-1] == 'YES':
                cell.fill = red

    # Sheet 2: Bottleneck Analysis
    ws2 = wb.create_sheet("Bottleneck Analysis")
    for col in 'ABCDEFG':
        ws2.column_dimensions[col].width = 22
    ws2['A1'] = "BOTTLENECK ANALYSIS — TOP 3 PROCESS CONSTRAINTS"
    ws2['A1'].font = Font(bold=True, size=14, color='1B2A4A')

    header(ws2, 3, ['Bottleneck', 'Steps Affected', 'Root Cause', 'Time Lost/Report', 'Monthly Impact (hrs)', 'Error Rate', 'Annual Cost'])
    bottlenecks = [
        ['Manual Data Entry', '2, 3, 4', 'No OCR/auto-import', '33 min', 231, '18%', '$124,740'],
        ['Approval Queue Delays', '6, 7, 8, 9', 'Email routing, no SLA', '72+ hrs avg', 840, '15% re-route', '$233,100'],
        ['Finance Reconciliation', '10, 11, 12', 'Manual validation, batch pay', '40 min + 5d', 560, '9.9% rejection', '$184,800'],
    ]
    for r, row_data in enumerate(bottlenecks, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = border
            cell.fill = red

    # Sheet 3: Benchmarking
    ws3 = wb.create_sheet("Benchmarking")
    for col in 'ABCDE':
        ws3.column_dimensions[col].width = 25
    ws3['A1'] = "INDUSTRY BENCHMARKING — EXPENSE PROCESSING"
    ws3['A1'].font = Font(bold=True, size=14, color='1B2A4A')

    header(ws3, 3, ['Metric', 'Current State', 'Industry Average', 'Best-in-Class', 'Gap'])
    benchmarks = [
        ['End-to-end cycle time', '9.7 days', '5 days', '2 days', '4.8x slower'],
        ['Cost per expense report', '$28.50', '$15.00', '$6.00', '4.75x higher'],
        ['Reports per FTE/day', '8', '15', '25', '3.1x below'],
        ['First-time approval rate', '82%', '90%', '95%', '13% gap'],
        ['Receipt capture rate', '81.5%', '90%', '98%', '16.5% gap'],
        ['Anomaly detection rate', '5% (manual)', '12%', '20%', 'Under-detected'],
        ['Employee satisfaction', 'Low', 'Medium', 'High', 'Significant gap'],
    ]
    for r, row_data in enumerate(benchmarks, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.border = border
        ws3.cell(row=r, column=5).fill = red

    wb.save(OUTPUT)
    print(f"Bottleneck table saved: {OUTPUT}")

if __name__ == "__main__":
    create()
